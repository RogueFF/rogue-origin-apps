from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Overfill Standards"

headers = ["Size Option", "Size (grams)", "Actual Weight (grams)", "Overfill (grams)", "Overfill %"]
ws.append(headers)

# (label, grams)
sizes = [
    ("1 oz",   28.3495),
    ("1/4 lb", 113.3981),
    ("1/2 lb", 226.7962),
    ("1 lb",   453.5924),
    ("10 lb",  4535.9237),
    ("5 kg",   5000.0),
]

for label, grams in sizes:
    ws.append([label, grams, None, None, None])

# Formulas for overfill columns (row 2 through 7)
for r in range(2, 2 + len(sizes)):
    ws[f"D{r}"] = f"=IF(C{r}=\"\",\"\",C{r}-B{r})"
    ws[f"E{r}"] = f"=IF(OR(C{r}=\"\",B{r}=0),\"\",(C{r}-B{r})/B{r})"

# Formatting
bold = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="668971")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
input_fill = PatternFill("solid", start_color="FFF8DC")

for col_idx, _ in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col_idx)
    c.font = bold
    c.fill = header_fill
    c.alignment = center
    c.border = border

for r in range(2, 2 + len(sizes)):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = Font(name="Arial")
        cell.border = border
        if col_idx in (2, 3, 4):
            cell.number_format = "#,##0.00"
        elif col_idx == 5:
            cell.number_format = "0.00%"
    # Highlight the input column
    ws.cell(row=r, column=3).fill = input_fill

widths = {"A": 14, "B": 16, "C": 24, "D": 18, "E": 14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"

# Note row
note_row = 2 + len(sizes) + 1
ws.cell(row=note_row, column=1, value="Note:").font = Font(name="Arial", bold=True)
ws.cell(row=note_row, column=2,
        value="Fill the yellow 'Actual Weight' column with the standardized overfilled weight for each size.")
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=5)

out = r"C:\Users\Koasm\Desktop\rogue-origin-apps\docs\overfill-standards.xlsx"
wb.save(out)
print(out)
